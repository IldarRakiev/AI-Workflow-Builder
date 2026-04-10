import base64
import logging
import tempfile
from pathlib import Path

from utils.llm import ask

logger = logging.getLogger(__name__)

MAX_DOC_CHARS = 0  # 0 = no limit


# ---------------------------------------------------------------------------
# Voice transcription (local faster-whisper)
# ---------------------------------------------------------------------------

_whisper_model = None


def _get_whisper_model():
    global _whisper_model
    if _whisper_model is None:
        from faster_whisper import WhisperModel
        logger.info("Loading Whisper model (base)...")
        _whisper_model = WhisperModel("base", device="cpu", compute_type="int8")
        logger.info("Whisper model loaded")
    return _whisper_model


async def transcribe_voice(file_path: str) -> str:
    """Transcribe an audio file (.ogg/.mp3/.wav) to text using local Whisper."""
    import asyncio

    def _sync_transcribe() -> str:
        model = _get_whisper_model()
        segments, info = model.transcribe(file_path, beam_size=5)
        text = " ".join(seg.text.strip() for seg in segments)
        logger.info("Transcribed %s (lang=%s, dur=%.1fs): %d chars",
                     file_path, info.language, info.duration, len(text))
        return text

    return await asyncio.get_event_loop().run_in_executor(None, _sync_transcribe)


# ---------------------------------------------------------------------------
# Document extraction (PDF / Excel / DOCX)
# ---------------------------------------------------------------------------

async def extract_document(file_path: str) -> str:
    """Extract text from PDF, Excel, or DOCX. Returns up to MAX_DOC_CHARS."""
    import asyncio

    path = Path(file_path)
    ext = path.suffix.lower()

    def _sync_extract() -> str:
        if ext == ".pdf":
            return _extract_pdf(path)
        elif ext in (".xlsx", ".xls"):
            return _extract_excel(path)
        elif ext in (".docx", ".doc"):
            return _extract_docx(path)
        else:
            return f"[Unsupported file format: {ext}]"

    text = await asyncio.get_event_loop().run_in_executor(None, _sync_extract)

    return text


def _extract_pdf(path: Path) -> str:
    from PyPDF2 import PdfReader
    reader = PdfReader(str(path))
    pages = []
    for i, page in enumerate(reader.pages):
        text = page.extract_text() or ""
        if text.strip():
            pages.append(f"--- Page {i + 1} ---\n{text.strip()}")
    return "\n\n".join(pages) if pages else "[PDF: no text extracted]"


def _extract_excel(path: Path) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(str(path), read_only=True, data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                rows.append("\t".join(cells))
        if rows:
            parts.append(f"--- Sheet: {sheet_name} ---\n" + "\n".join(rows))
    wb.close()
    return "\n\n".join(parts) if parts else "[Excel: no data extracted]"


def _extract_docx(path: Path) -> str:
    from docx import Document
    doc = Document(str(path))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n\n".join(paragraphs) if paragraphs else "[DOCX: no text extracted]"


# ---------------------------------------------------------------------------
# Image description (vision via OpenRouter)
# ---------------------------------------------------------------------------

async def describe_image(file_path: str, ask_kwargs: dict = {}) -> str:
    """Describe an image by sending it as base64 to a vision-capable model."""
    path = Path(file_path)
    suffix = path.suffix.lower().lstrip(".")
    mime_map = {"jpg": "jpeg", "jpeg": "jpeg", "png": "png", "gif": "gif", "webp": "webp"}
    mime_type = f"image/{mime_map.get(suffix, 'jpeg')}"

    with open(path, "rb") as f:
        b64 = base64.b64encode(f.read()).decode("utf-8")

    messages = [
        {
            "role": "user",
            "content": [
                {
                    "type": "image_url",
                    "image_url": {"url": f"data:{mime_type};base64,{b64}"},
                },
                {
                    "type": "text",
                    "text": "Describe this image in detail. Write in the same language the user has been using (default: Russian).",
                },
            ],
        }
    ]

    return await ask(messages, **ask_kwargs)
